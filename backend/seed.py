"""Mode « seed URLs » (Phase 7) : ingestion directe d'URLs connues.

Permet de reproduire la couverture d'une base externe (p. ex. l'Excel de
référence) sans dépendre d'un moteur de recherche : on lit une liste d'URLs
(`.txt`, `.csv` ou `.xlsx`, typiquement la colonne « Lien source »), on construit
des articles `{titre, texte, url, date, source}`, puis on les fait passer dans la
chaîne standard :

    fetch_text(url) -> extraction IA -> géocodage (avec repli lieu-dit) ->
    normalisation commune administrative -> validation -> upsert closure/source.

Contrairement au pipeline de collecte, le préfiltre n'est PAS appliqué : ces URLs
sont curées et explicitement fournies. Aucune dépendance réseau dans les tests :
`fetch_fn`, `extractor_fn` et `geocode_fn` sont injectables.
"""
from __future__ import annotations

import csv
import re
import unicodedata
from datetime import date, timedelta
from pathlib import Path

from backend import commune_normalize, ingest_map, store, validation
from backend.dedup import closure_id
from backend.extractor import normalise_banque

_URL_RE = re.compile(r"https?://\S+")

# En-têtes reconnus (normalisés : minuscules, sans accents, séparateurs -> espace).
_URL_HEADERS = {"url", "lien source", "lien", "lien_source", "source url"}
_BANQUE_HEADERS = {"banque"}
_COMMUNE_HEADERS = {"commune"}
_LOC_HEADERS = {"agence localisation", "agence / localisation", "localisation"}
_DATE_HEADERS = {"date", "date de fermeture", "date_fermeture"}
_SOURCE_HEADERS = {"source"}
_DEPT_HEADERS = {"departement"}
_ELEMENTS_HEADERS = {"elements retenus", "elements"}

_DEPARTEMENT_CODES = {
    "ain": "01", "aisne": "02", "allier": "03", "alpes de haute provence": "04",
    "hautes alpes": "05", "alpes maritimes": "06", "ardeche": "07", "ardennes": "08",
    "ariege": "09", "aube": "10", "aude": "11", "aveyron": "12",
    "bouches du rhone": "13", "calvados": "14", "cantal": "15", "charente": "16",
    "charente maritime": "17", "cher": "18", "correze": "19", "corse du sud": "2A",
    "haute corse": "2B", "cote d or": "21", "cotes d armor": "22", "creuse": "23",
    "dordogne": "24", "doubs": "25", "drome": "26", "eure": "27",
    "eure et loir": "28", "finistere": "29", "gard": "30", "haute garonne": "31",
    "gers": "32", "gironde": "33", "herault": "34", "ille et vilaine": "35",
    "indre": "36", "indre et loire": "37", "isere": "38", "jura": "39",
    "landes": "40", "loir et cher": "41", "loire": "42", "haute loire": "43",
    "loire atlantique": "44", "loiret": "45", "lot": "46", "lot et garonne": "47",
    "lozere": "48", "maine et loire": "49", "manche": "50", "marne": "51",
    "haute marne": "52", "mayenne": "53", "meurthe et moselle": "54", "meuse": "55",
    "morbihan": "56", "moselle": "57", "nievre": "58", "nord": "59",
    "oise": "60", "orne": "61", "pas de calais": "62", "puy de dome": "63",
    "pyrenees atlantiques": "64", "hautes pyrenees": "65",
    "pyrenees orientales": "66", "bas rhin": "67", "haut rhin": "68",
    "rhone": "69", "haute saone": "70", "saone et loire": "71", "sarthe": "72",
    "savoie": "73", "haute savoie": "74", "paris": "75", "seine maritime": "76",
    "seine et marne": "77", "yvelines": "78", "deux sevres": "79", "somme": "80",
    "tarn": "81", "tarn et garonne": "82", "var": "83", "vaucluse": "84",
    "vendee": "85", "vienne": "86", "haute vienne": "87", "vosges": "88",
    "yonne": "89", "territoire de belfort": "90", "essonne": "91",
    "hauts de seine": "92", "seine saint denis": "93", "val de marne": "94",
    "val d oise": "95",
}


def _norm_header(key: str | None) -> str:
    sans = "".join(
        c for c in unicodedata.normalize("NFD", key or "")
        if unicodedata.category(c) != "Mn"
    )
    return re.sub(r"[^a-z0-9]+", " ", sans.lower()).strip()


def _first(mapping: dict, headers: set[str]) -> str:
    for k, v in mapping.items():
        if _norm_header(k) in headers:
            return "" if v is None else str(v).strip()
    return ""


def _dept_code(value: str) -> str:
    value = (value or "").strip()
    if value in _DEPARTEMENT_CODES.values():
        return value
    return _DEPARTEMENT_CODES.get(_norm_header(value), value)


def _date_fermeture(value: str | None) -> str | None:
    value = (value or "").strip()
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", value):
        return value
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", value)
    if m:
        jour, mois, annee = map(int, m.groups())
        try:
            d = date(annee, mois, jour)
        except ValueError:
            return value or None
        if "semaine" in _norm_header(value) and "preced" in _norm_header(value):
            d = d - timedelta(days=7)
        return d.isoformat()
    return value or None


def _row_to_article(mapping: dict) -> dict | None:
    """Construit un article depuis une ligne (csv/xlsx) ; None si pas d'URL."""
    url_cell = _first(mapping, _URL_HEADERS)
    m = _URL_RE.search(url_cell)
    url = m.group(0).rstrip('.,;)') if m else ""
    if not url:
        return None
    banque = _first(mapping, _BANQUE_HEADERS)
    commune = _first(mapping, _COMMUNE_HEADERS)
    loc = _first(mapping, _LOC_HEADERS)
    elements = _first(mapping, _ELEMENTS_HEADERS)
    departement = _dept_code(_first(mapping, _DEPT_HEADERS))
    # Titre = signal fort pour l'extraction (banque + localisation + fermeture).
    titre_parts = [p for p in (banque, loc or commune) if p]
    titre = " ".join(titre_parts) + " fermeture agence" if titre_parts else url
    return {
        "titre": titre.strip(),
        "texte": elements or "",
        "url": url,
        "date": _first(mapping, _DATE_HEADERS) or None,
        "source": _first(mapping, _SOURCE_HEADERS) or "Seed URL",
        "departement": departement or None,
        "commune_attendue": commune or None,
        "agence_localisation": loc or None,
        "seed_communes": [commune] if commune else [],
        "seed_targets": [{
            "banque": banque,
            "commune": commune,
            "departement": departement or None,
            "date_fermeture": _date_fermeture(_first(mapping, _DATE_HEADERS)),
            "agence_localisation": loc or None,
        }] if commune else [],
    }


def _article_from_url(url: str) -> dict:
    return {"titre": url, "texte": "", "url": url, "date": None,
            "source": "Seed URL", "departement": None}


def load_articles(path) -> list[dict]:
    """Charge une liste d'articles seed depuis .txt / .csv / .xlsx."""
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm", ".xls"):
        return _load_xlsx(path)
    if suffix == ".csv":
        return _load_csv(path)
    return _load_txt(path)


def _load_txt(path: Path) -> list[dict]:
    out: list[dict] = []
    seen: set[str] = set()
    for line in path.read_text(encoding="utf-8").splitlines():
        m = _URL_RE.search(line.strip())
        if not m:
            continue
        url = m.group(0).rstrip('.,;)')
        if url not in seen:
            seen.add(url)
            out.append(_article_from_url(url))
    return out


def _load_csv(path: Path) -> list[dict]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return _articles_from_rows(reader)


def _load_xlsx(path: Path) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
    except StopIteration:
        return []
    mappings = []
    for values in rows_iter:
        if values is None or all(v is None or str(v).strip() == "" for v in values):
            continue
        mappings.append({header[i]: values[i] if i < len(values) else ""
                         for i in range(len(header))})
    return _articles_from_rows(mappings)


def _articles_from_rows(rows) -> list[dict]:
    out: list[dict] = []
    by_url: dict[str, dict] = {}
    for mapping in rows:
        art = _row_to_article(mapping)
        if not art:
            continue
        existing = by_url.get(art["url"])
        if existing is None:
            by_url[art["url"]] = art
            out.append(art)
            continue
        for field in ("titre", "texte"):
            value = art.get(field) or ""
            if value and value not in (existing.get(field) or ""):
                existing[field] = f"{existing.get(field) or ''}\n{value}".strip()
        communes = existing.setdefault("seed_communes", [])
        commune = art.get("commune_attendue")
        if commune and commune not in communes:
            communes.append(commune)
        targets = existing.setdefault("seed_targets", [])
        for target in art.get("seed_targets") or []:
            if target.get("commune") and target not in targets:
                targets.append(target)
    for art in out:
        communes = art.get("seed_communes") or []
        if len(communes) > 1:
            liste = ", ".join(communes[:-1]) + f" et {communes[-1]}"
            art["texte"] = (
                f"{art.get('texte') or ''}\n"
                f"Cet article concerne plusieurs agences. "
                f"Sont concernées les agences de {liste}."
            ).strip()
    return out


def _closures_depuis_targets(article: dict, geocode_fn, *, allow_single: bool = False) -> list[dict]:
    targets = article.get("seed_targets") or []
    if len(targets) <= 1 and not allow_single:
        return []
    out: list[dict] = []
    for target in targets:
        banque = normalise_banque(target.get("banque") or "")
        commune = target.get("commune")
        if not banque or not commune:
            continue
        geo = geocode_fn(commune, target.get("departement")) or {}
        agence_loc = target.get("agence_localisation")
        closure = {
            "id": closure_id(banque, commune, "fermeture", adresse=agence_loc),
            "banque": banque,
            "commune": geo.get("commune") or commune,
            "code_insee": geo.get("code_insee"),
            "departement": geo.get("departement") or target.get("departement"),
            "type": "fermeture",
            "date_annonce": article.get("date") or None,
            "date_fermeture": target.get("date_fermeture") or article.get("date"),
            "statut": "confirmé",
            "statut_temporel": "a_venir",
            "date_fermeture_approx": 0,
            "fiabilite": 4,
            "lat": geo.get("lat"),
            "lon": geo.get("lon"),
            "citation": article.get("titre") or article.get("texte") or article.get("url"),
            "agence_localisation": (
                agence_loc if agence_loc and agence_loc != (geo.get("commune") or commune)
                else commune if geo.get("commune") and geo.get("commune") != commune
                else None
            ),
            "commune_originale": (
                agence_loc if agence_loc and agence_loc != (geo.get("commune") or commune)
                else commune if geo.get("commune") and geo.get("commune") != commune
                else None
            ),
        }
        out.append(closure)
    return out


def _persist_unlocated(conn, closure: dict, article: dict, url: str, raison: str) -> None:
    store.upsert_closure_unlocated(conn, {
        "id": closure_id(
            closure.get("banque") or "",
            closure.get("commune") or "",
            closure.get("type") or "fermeture",
            adresse=f"{url}|{closure.get('agence_localisation') or ''}",
        ),
        "banque": closure.get("banque"),
        "commune": closure.get("commune"),
        "departement": closure.get("departement"),
        "type": closure.get("type"),
        "date_fermeture": closure.get("date_fermeture"),
        "statut": closure.get("statut"),
        "statut_temporel": closure.get("statut_temporel"),
        "fiabilite": closure.get("fiabilite"),
        "citation": closure.get("citation"),
        "url": url or None,
        "titre": article.get("titre"),
        "source": article.get("source"),
        "date": article.get("date"),
        "raison": raison,
    })


def _publish_or_unlocated(conn, closure: dict, article: dict, url: str, geocode_fn) -> bool:
    try:
        geo = geocode_fn(closure["commune"], closure.get("departement"))
    except Exception:
        geo = None
    if geo:
        closure["lat"] = closure.get("lat") or geo.get("lat")
        closure["lon"] = closure.get("lon") or geo.get("lon")
        if not validation.departement_valide(closure.get("departement")):
            closure["departement"] = geo.get("departement")
        if not closure.get("code_insee"):
            closure["code_insee"] = geo.get("code_insee")
        closure = commune_normalize.appliquer(closure, geo)
    publiable, raison = validation.fermeture_publiable(closure, geo)
    if not publiable:
        _persist_unlocated(conn, closure, article, url, raison)
        return False
    store.upsert_closure(conn, closure)
    store.add_source(conn, closure["id"], {
        "url": url, "titre": article.get("titre"),
        "source": article.get("source"), "date": article.get("date"),
    })
    return True


def _persist_structured_signals(conn, result: dict, article: dict, url: str) -> int:
    count = 0
    for signal in result.get("department_signals") or []:
        banque = normalise_banque(signal.get("bank") or "") if signal.get("bank") else None
        store.upsert_department_signal(conn, {
            "id": closure_id(
                banque or "",
                signal.get("departement") or "",
                "department_signal",
                adresse=f"{url}|{signal.get('count') or ''}",
            ),
            "banque": banque,
            "departement": signal.get("departement") or article.get("departement"),
            "count": signal.get("count"),
            "communes_mentioned": ", ".join(signal.get("communes_mentioned") or []),
            "confidence": signal.get("confidence"),
            "evidence": signal.get("evidence"),
            "url": url or None,
            "titre": article.get("titre"),
            "source": article.get("source"),
            "date": article.get("date"),
        })
        count += 1
    for signal in result.get("vague_signals") or []:
        banque = normalise_banque(signal.get("bank") or "") if signal.get("bank") else None
        store.upsert_vague_signal(conn, {
            "id": closure_id(
                banque or "",
                signal.get("scope") or "",
                "vague_signal",
                adresse=f"{url}|{signal.get('count') or ''}",
            ),
            "banque": banque,
            "scope": signal.get("scope"),
            "count": signal.get("count"),
            "confidence": signal.get("confidence"),
            "evidence": signal.get("evidence"),
            "url": url or None,
            "titre": article.get("titre"),
            "source": article.get("source"),
            "date": article.get("date"),
        })
        count += 1
    return count


def _is_structured_result(resultat: dict) -> bool:
    return any(k in resultat for k in ("closures", "department_signals", "vague_signals", "article_type"))


def _ingest_structured_result(conn, resultat: dict, article: dict, url: str, geocode_fn) -> tuple[int, int, int]:
    """Stocke un résultat ExtractionResult : carte stricte, base large."""
    signals = _persist_structured_signals(conn, resultat, article, url)
    closures_map, vigilance = ingest_map.map_result(resultat, article, date.today().isoformat())
    published = rejected = 0
    for closure in closures_map:
        if _publish_or_unlocated(conn, closure, article, url, geocode_fn):
            published += 1
        else:
            rejected += 1
    if vigilance:
        store.upsert_vigilance(conn, vigilance)
        signals += 1
    return published, rejected, signals


def ingest(
    conn,
    articles: list[dict],
    *,
    extractor_fn,
    geocode_fn,
    fetch_fn=None,
    since_date: str | None = None,
) -> dict:
    """Ingère des articles seed (URLs curées) → fermetures publiées.

    Best-effort par article : une erreur d'extraction n'interrompt pas le lot.
    Retourne un récapitulatif {urls, extraits, fermetures, rejets, vigilances}.
    """
    recap = {"urls": 0, "extraits": 0, "fermetures": 0, "rejets": 0, "vigilances": 0}
    for art in articles:
        recap["urls"] += 1
        url = art.get("url") or ""
        texte = art.get("texte") or ""
        # Enrichissement plein texte si le contenu fourni est trop maigre.
        if fetch_fn and url and len(texte) < 400:
            try:
                complet = fetch_fn(url)
                if complet:
                    art["texte"] = (texte + "\n\n" + complet)[:6000]
            except Exception as exc:
                print(f"[seed] fetch en erreur ({url}): {exc}")
        plan_closures = _closures_depuis_targets(art, geocode_fn)
        if plan_closures:
            for closure in plan_closures:
                if _publish_or_unlocated(conn, closure, art, url, geocode_fn):
                    recap["fermetures"] += 1
                else:
                    recap["rejets"] += 1
            if url:
                store.mark_url_seen(conn, url)
            continue
        try:
            resultat = extractor_fn(art)
        except Exception as exc:
            print(f"[seed] extraction en erreur ({url}): {exc}")
            fallback = _closures_depuis_targets(art, geocode_fn, allow_single=True)
            for closure in fallback:
                if _publish_or_unlocated(conn, closure, art, url, geocode_fn):
                    recap["fermetures"] += 1
                else:
                    recap["rejets"] += 1
            if fallback and url:
                store.mark_url_seen(conn, url)
            continue
        if url:
            store.mark_url_seen(conn, url)
        if resultat is None:
            fallback = _closures_depuis_targets(art, geocode_fn, allow_single=True)
            for closure in fallback:
                if _publish_or_unlocated(conn, closure, art, url, geocode_fn):
                    recap["fermetures"] += 1
                else:
                    recap["rejets"] += 1
            if fallback:
                continue
            recap["vigilances"] += 1
            continue
        recap["extraits"] += 1
        if isinstance(resultat, dict) and _is_structured_result(resultat):
            published, rejected, signals = _ingest_structured_result(
                conn, resultat, art, url, geocode_fn)
            recap["fermetures"] += published
            recap["rejets"] += rejected
            recap["vigilances"] += signals
            continue
        try:
            expected_commune = art.get("commune_attendue")
            if expected_commune:
                original = resultat.get("commune")
                if original and original != expected_commune:
                    resultat["agence_localisation"] = art.get("agence_localisation") or original
                    resultat["commune_originale"] = original
                resultat["commune"] = expected_commune
            geo = geocode_fn(resultat["commune"], resultat.get("departement"))
        except Exception:
            geo = None
        if geo:
            resultat["lat"] = resultat.get("lat") or geo.get("lat")
            resultat["lon"] = resultat.get("lon") or geo.get("lon")
            if not validation.departement_valide(resultat.get("departement")):
                resultat["departement"] = geo.get("departement")
            if not resultat.get("code_insee"):
                resultat["code_insee"] = geo.get("code_insee")
            resultat = commune_normalize.appliquer(resultat, geo)
        publiable, raison = validation.fermeture_publiable(resultat, geo)
        if not publiable:
            _persist_unlocated(conn, resultat, art, url, raison)
            recap["rejets"] += 1
            continue
        store.upsert_closure(conn, resultat)
        store.add_source(conn, resultat["id"], {
            "url": url, "titre": art.get("titre"),
            "source": art.get("source"), "date": art.get("date"),
        })
        recap["fermetures"] += 1
    return recap
