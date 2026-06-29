"""Benchmark de couverture Copilot (Cycle 1, read-only).

Classe chaque ligne du fichier de référence Copilot sur deux axes :
  - couverture dans notre base (data.json) : present_on_map / present_unlocated /
    present_department / needs_research / rejected_with_reason / confirmed_missing ;
  - fiabilité de la source Copilot : high / medium / low (+ source_flag).

Ne modifie jamais le pipeline. Produit data/export/copilot_coverage.{csv,json}.
"""
from __future__ import annotations

import json
import math
from pathlib import Path

from backend.dedup import normalise_cle
from backend.drilldown import est_plan
from tools.compare_expected_closures import _cle_banque, _cle_commune
import config

# Mapping positionnel des colonnes de l'Excel Copilot (en-tête banque = "²").
COPILOT_COLS: dict[str, int] = {
    "banque": 0, "agence_localisation": 1, "adresse": 2, "commune": 3,
    "departement": 4, "region": 5, "lat": 6, "lon": 7, "date_fermeture": 8,
    "precision_date": 9, "source": 10, "url": 11, "score": 14,
    "statut_copilot": 15, "commentaires": 16,
}
_RAW_COLS = {"lat", "lon"}  # conservés bruts (float), pas de .strip()


def _cell(values, idx):
    return values[idx] if idx < len(values) else None


def load_copilot_rows(path) -> list[dict]:
    """Charge l'Excel Copilot en liste de dicts (mapping positionnel)."""
    from openpyxl import load_workbook

    wb = load_workbook(Path(path), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    next(rows_iter, None)  # saute l'en-tête
    out: list[dict] = []
    for values in rows_iter:
        if values is None or all(v is None or str(v).strip() == "" for v in values):
            continue
        row: dict = {}
        for champ, idx in COPILOT_COLS.items():
            v = _cell(values, idx)
            if champ in _RAW_COLS:
                row[champ] = v if v is not None else ""
            else:
                row[champ] = "" if v is None else str(v).strip()
        out.append(row)
    return out


# Table inverse {nom normalisé du département -> code}.
_DEPT_NAME_TO_CODE = {normalise_cle(nom): code for code, nom in config.DEPARTEMENTS.items()}


def _norm(s) -> str:
    return normalise_cle(s or "")


def dept_name_to_code(name) -> str | None:
    return _DEPT_NAME_TO_CODE.get(normalise_cle(name or "")) or None


def load_overrides(path) -> dict:
    """Charge tools/copilot_overrides.json. Fichier absent/None -> sections vides."""
    if not path or not Path(path).exists():
        return {"sources": [], "rows": []}
    data = json.loads(Path(path).read_text(encoding="utf-8"))
    return {"sources": data.get("sources") or [], "rows": data.get("rows") or []}


def _haversine_m(lat1, lon1, lat2, lon2) -> float:
    r = 6371000.0
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (math.sin(dlat / 2) ** 2
         + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon / 2) ** 2)
    return 2 * r * math.asin(math.sqrt(a))


def _row_matches_closure(banque_cle: str, commune_cle: str, cl: dict) -> bool:
    if not commune_cle or _cle_banque(cl.get("banque")) != banque_cle:
        return False
    for champ in ("commune", "agence_localisation", "commune_originale"):
        if _cle_commune(cl.get(champ)) == commune_cle:
            return True
    return False


def _as_float(v):
    try:
        if v in (None, ""):
            return None
        return float(v)
    except (TypeError, ValueError):
        return None


def _has_department_signal(banque_cle: str, dept_code: str, payload: dict) -> bool:
    for v in payload.get("vigilances") or []:
        if v.get("departement") == dept_code and _cle_banque(v.get("banque")) == banque_cle:
            return True
    est = (payload.get("department_estimates") or {}).get(dept_code)
    if est:
        for sig in est.get("signals") or []:
            if _cle_banque(sig.get("banque")) == banque_cle:
                return True
    return False


def classify_coverage(row: dict, payload: dict) -> dict:
    banque_cle = _cle_banque(row.get("banque"))
    commune_cle = _cle_commune(row.get("commune"))
    for cl in payload.get("closures") or []:
        if not _row_matches_closure(banque_cle, commune_cle, cl):
            continue
        cl_lat, cl_lon = _as_float(cl.get("lat")), _as_float(cl.get("lon"))
        has_geo = cl_lat is not None and cl_lon is not None
        statut = cl.get("statut") or cl.get("statut_temporel") or ""
        if not has_geo:
            return {"status": "present_unlocated", "match_type": "commune",
                    "pipeline_id": cl.get("id", ""), "pipeline_status": statut}
        match_type = "commune"
        row_lat, row_lon = _as_float(row.get("lat")), _as_float(row.get("lon"))
        if row_lat is not None and row_lon is not None:
            if _haversine_m(row_lat, row_lon, cl_lat, cl_lon) < 500:
                match_type = "exact"
        return {"status": "present_on_map", "match_type": match_type,
                "pipeline_id": cl.get("id", ""), "pipeline_status": statut}

    dept_code = dept_name_to_code(row.get("departement"))
    if dept_code and _has_department_signal(banque_cle, dept_code, payload):
        return {"status": "present_department", "match_type": "département",
                "pipeline_id": "", "pipeline_status": ""}
    return {"status": "needs_research", "match_type": "aucun",
            "pipeline_id": "", "pipeline_status": ""}


_ROW_OVERRIDE_FIELDS = ("status", "missing_reason", "next_action",
                        "source_reliability", "source_flag")


def apply_reliability(row: dict, overrides: dict) -> dict:
    result: dict = {}
    src = _norm(row.get("source"))
    has_url = bool((row.get("url") or "").strip())

    # 1. Règles par motif de source.
    for rule in overrides.get("sources") or []:
        if _norm(rule.get("match_source")) not in src:
            continue
        if rule.get("require_no_url") and has_url:
            continue
        for k in ("source_reliability", "source_flag"):
            if rule.get(k):
                result[k] = rule[k]
        if rule.get("default_status_if_uncovered"):
            result["_source_default_status"] = rule["default_status_if_uncovered"]
        if rule.get("default_next_action"):
            result["_source_default_next_action"] = rule["default_next_action"]

    # 2. Règles par ligne (priment sur les règles de source).
    banque_cle = _cle_banque(row.get("banque"))
    commune_cle = _cle_commune(row.get("commune"))
    for rule in overrides.get("rows") or []:
        m = rule.get("match") or {}
        if _cle_banque(m.get("banque")) != banque_cle or _cle_commune(m.get("commune")) != commune_cle:
            continue
        al = m.get("agence_localisation")
        if al and _cle_commune(al) != _cle_commune(row.get("agence_localisation")):
            continue
        for k in _ROW_OVERRIDE_FIELDS:
            if rule.get(k) is not None:
                result[k] = rule[k]

    # 3. Heuristique de fiabilité par défaut.
    if not result.get("source_reliability"):
        result["source_reliability"] = "medium" if has_url else "low"
    return result


COVERAGE_STATUSES = ["present_on_map", "present_unlocated", "present_department",
                     "needs_research", "rejected_with_reason", "confirmed_missing"]

RECORD_FIELDS = [
    "banque", "agence_localisation", "commune", "departement", "adresse",
    "lat", "lon", "source", "url", "score_copilot", "statut_copilot",
    "matched_pipeline", "match_type", "pipeline_id", "pipeline_status",
    "status", "missing_reason", "next_action", "source_reliability", "source_flag",
]

_NEXT_ACTION_BY_STATUS = {
    "present_on_map": "Aucune — déjà sur la carte ; contrôler la fiabilité de la source.",
    "present_unlocated": "Géocoder à l'adresse précise pour publication carte.",
    "present_department": "Identifier l'agence/commune précise pour faire monter le signal départemental.",
    "rejected_with_reason": "Voir missing_reason ; ne pas publier.",
    "confirmed_missing": "Voir missing_reason ; intégrer si une source fiable est retrouvée.",
}


def default_next_action_queries(row: dict) -> str:
    banque, commune = row.get("banque", ""), row.get("commune", "")
    requetes = [
        f'"{banque}" "{commune}" "fermeture agence"',
        f'"{banque}" "{commune}" "agence ferme"',
        f'"{banque}" "{commune}" "regroupement agence"',
        f'"{commune}" "banque ferme"',
    ]
    contexte = " ".join(filter(None, [row.get("agence_localisation"),
                                      row.get("statut_copilot"), row.get("commentaires")]))
    if est_plan(contexte):
        requetes.append(f'"{banque}" "plan" "fermeture" "agences"')
    return " | ".join(requetes[:5])


def build_record(row: dict, payload: dict, overrides: dict) -> dict:
    cov = classify_coverage(row, payload)
    rel = apply_reliability(row, overrides)

    status = rel.get("status")
    if not status:
        if cov["status"] == "needs_research" and rel.get("_source_default_status"):
            status = rel["_source_default_status"]
        else:
            status = cov["status"]

    next_action = rel.get("next_action") or rel.get("_source_default_next_action")
    if not next_action:
        next_action = (default_next_action_queries(row) if status == "needs_research"
                       else _NEXT_ACTION_BY_STATUS.get(status, "À qualifier."))

    return {
        "banque": row.get("banque", ""),
        "agence_localisation": row.get("agence_localisation", ""),
        "commune": row.get("commune", ""),
        "departement": row.get("departement", ""),
        "adresse": row.get("adresse", ""),
        "lat": row.get("lat", ""),
        "lon": row.get("lon", ""),
        "source": row.get("source", ""),
        "url": row.get("url", ""),
        "score_copilot": row.get("score", ""),
        "statut_copilot": row.get("statut_copilot", ""),
        "matched_pipeline": "oui" if cov["match_type"] != "aucun" else "non",
        "match_type": cov["match_type"],
        "pipeline_id": cov["pipeline_id"],
        "pipeline_status": cov["pipeline_status"],
        "status": status,
        "missing_reason": rel.get("missing_reason", ""),
        "next_action": next_action,
        "source_reliability": rel["source_reliability"],
        "source_flag": rel.get("source_flag", ""),
    }
