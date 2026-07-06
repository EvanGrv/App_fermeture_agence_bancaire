"""Diagnostic automatisé de couverture.

Compare un fichier de référence (Excel ou CSV) listant les fermetures attendues
au fichier `data/export/data.json` produit par le pipeline, et classe chaque
ligne attendue dans l'une des catégories :

    present_closure              -> fermeture publiée, complète
    missing_date                 -> fermeture publiée mais sans date_fermeture
    present_malformed            -> fermeture publiée mais incomplète (pas d'INSEE)
    bad_commune_normalization    -> captée sous la localisation d'agence au lieu de
                                    la commune administrative
    present_vigilance            -> absente des fermetures, présente en vigilance
    plan_not_exploded            -> plan multi-agences capté mais commune non éclatée
    absent                       -> introuvable

Usage CLI :
    python -m tools.compare_expected_closures reference.xlsx data/export/data.json
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import unicodedata
from pathlib import Path

from backend.dedup import normalise_cle
from backend.drilldown import est_plan
from backend.extractor import normalise_banque

STATUS_PRESENT = "present_closure"
STATUS_MISSING_DATE = "missing_date"
STATUS_MALFORMED = "present_malformed"
STATUS_BAD_COMMUNE = "bad_commune_normalization"
STATUS_VIGILANCE = "present_vigilance"
STATUS_PLAN = "plan_not_exploded"
STATUS_ABSENT = "absent"

ALL_STATUSES = [
    STATUS_PRESENT, STATUS_MISSING_DATE, STATUS_MALFORMED, STATUS_BAD_COMMUNE,
    STATUS_VIGILANCE, STATUS_PLAN, STATUS_ABSENT,
]

_TRUE = {"1", "true", "vrai", "oui", "yes", "x"}


# --- Normalisation ----------------------------------------------------------

def _cle_commune(commune: str | None) -> str:
    """Clé de commune insensible aux tirets/apostrophes (cf. validation._cle_commune)."""
    return re.sub(r"[-'’\s]+", " ", normalise_cle(commune or "")).strip()


def _cle_banque(banque: str | None) -> str:
    cle = normalise_cle(banque or "")
    if "credit municipal" in cle:
        return "credit municipal"
    return normalise_cle(normalise_banque(banque or ""))


def _bool(value) -> bool:
    return str(value or "").strip().lower() in _TRUE


# --- Chargement des entrées -------------------------------------------------

# En-têtes reconnus (normalisés : minuscules, sans accents, séparateurs -> espace).
# Couvre l'Excel français de référence ET le CSV interne minimal.
_HEADER_ALIASES = {
    "banque": "banque",
    "commune": "commune",
    "date fermeture": "date_fermeture",
    "date de fermeture": "date_fermeture",
    "agence localisation": "agence_localisation",
    "agence localisation commune": "agence_localisation",
    "plan": "plan",
    "departement": "departement",
    "region": "region",
    "precision de date": "precision",
    "elements retenus": "elements",
    "lien source": "lien_source",
    "score de confiance": "score",
}


def _norm_header(key: str | None) -> str:
    sans = "".join(
        c for c in unicodedata.normalize("NFD", key or "")
        if unicodedata.category(c) != "Mn"
    )
    return re.sub(r"[^a-z0-9]+", " ", sans.lower()).strip()


def _row_from_mapping(raw: dict) -> dict:
    canon: dict[str, str] = {}
    for k, v in raw.items():
        champ = _HEADER_ALIASES.get(_norm_header(k))
        if champ and champ not in canon:
            canon[champ] = "" if v is None else str(v).strip()
    if not canon.get("banque") and raw:
        # L'Excel Copilot historique utilise "²" comme en-tête de banque.
        first_value = next(iter(raw.values()))
        canon["banque"] = "" if first_value is None else str(first_value).strip()

    # Le plan multi-agences n'a pas de colonne dédiée dans l'Excel : on le déduit
    # des champs textuels ("dix agences", "plan de fermeture"...).
    if "plan" in canon:
        plan = _bool(canon.get("plan"))
    else:
        contexte = " ".join(
            canon.get(f, "") for f in ("agence_localisation", "precision", "elements")
        )
        plan = est_plan(contexte)

    return {
        "banque": canon.get("banque", ""),
        "commune": canon.get("commune", ""),
        "date_fermeture": canon.get("date_fermeture", ""),
        "agence_localisation": canon.get("agence_localisation", ""),
        "plan": plan,
    }


def load_expected(path) -> list[dict]:
    """Charge le fichier de référence (.csv ou .xlsx) en liste de dicts normalisés."""
    path = Path(path)
    if path.suffix.lower() in (".xlsx", ".xlsm", ".xls"):
        return _load_xlsx(path)
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return [_row_from_mapping(r) for r in reader if any((v or "").strip() for v in r.values())]


def _load_xlsx(path: Path) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
    except StopIteration:
        return []
    out = []
    for values in rows_iter:
        if values is None or all(v is None or str(v).strip() == "" for v in values):
            continue
        raw = {header[i]: values[i] if i < len(values) else "" for i in range(len(header))}
        out.append(_row_from_mapping(raw))
    return out


def load_payload(path) -> dict:
    return json.loads(Path(path).read_text(encoding="utf-8"))


# --- Classification ---------------------------------------------------------

def _matches(banque_cle: str, commune_cle: str, item: dict, item_commune_field: str) -> bool:
    return (
        _cle_banque(item.get("banque")) == banque_cle
        and _cle_commune(item.get(item_commune_field)) == commune_cle
    )


def classify(expected: dict, payload: dict) -> str:
    banque_cle = _cle_banque(expected.get("banque"))
    commune_cle = _cle_commune(expected.get("commune"))
    loc_cle = _cle_commune(expected.get("agence_localisation"))
    closures = payload.get("closures") or []
    vigilances = payload.get("vigilances") or []

    # 1. Fermeture publiée sur la commune administrative attendue.
    for cl in closures:
        if _matches(banque_cle, commune_cle, cl, "commune"):
            if not cl.get("code_insee"):
                return STATUS_MALFORMED
            if expected.get("date_fermeture") and not cl.get("date_fermeture"):
                return STATUS_MISSING_DATE
            return STATUS_PRESENT
        if _matches(banque_cle, commune_cle, cl, "agence_localisation"):
            if not cl.get("code_insee"):
                return STATUS_MALFORMED
            if expected.get("date_fermeture") and not cl.get("date_fermeture"):
                return STATUS_MISSING_DATE
            return STATUS_PRESENT

    # 2. Captée sous la localisation d'agence au lieu de la commune administrative.
    if loc_cle:
        for cl in closures:
            if _matches(banque_cle, loc_cle, cl, "commune"):
                return STATUS_BAD_COMMUNE

    # 3. Plan multi-agences capté en vigilance mais commune non éclatée.
    if expected.get("plan"):
        for v in vigilances:
            if _cle_banque(v.get("banque")) == banque_cle and est_plan(
                f"{v.get('titre','')} {v.get('extrait','')}"
            ):
                return STATUS_PLAN

    # 4. Présente en vigilance (commune ou localisation citée dans le texte).
    for v in vigilances:
        if _cle_banque(v.get("banque")) != banque_cle:
            continue
        texte = _cle_commune(f"{v.get('titre','')} {v.get('extrait','')}")
        if commune_cle and commune_cle in texte:
            return STATUS_VIGILANCE
        if loc_cle and loc_cle in texte:
            return STATUS_VIGILANCE

    return STATUS_ABSENT


def compare(rows: list[dict], payload: dict) -> list[dict]:
    return [{**row, "status": classify(row, payload)} for row in rows]


def summarize(results: list[dict]) -> dict[str, int]:
    summary = {status: 0 for status in ALL_STATUSES}
    for r in results:
        summary[r["status"]] = summary.get(r["status"], 0) + 1
    return summary


# --- CLI --------------------------------------------------------------------

def main(argv=None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("reference", help="Fichier de référence (.xlsx ou .csv)")
    parser.add_argument("payload", help="Chemin vers data.json")
    parser.add_argument("--only", choices=ALL_STATUSES, help="Ne lister qu'un statut")
    args = parser.parse_args(argv)

    rows = load_expected(args.reference)
    payload = load_payload(args.payload)
    results = compare(rows, payload)
    summary = summarize(results)

    for r in results:
        if args.only and r["status"] != args.only:
            continue
        print(f"{r['status']:<28} {r['banque']} — {r['commune']}")
    print("\n--- Récapitulatif ---")
    for status in ALL_STATUSES:
        print(f"{status:<28} {summary[status]}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
